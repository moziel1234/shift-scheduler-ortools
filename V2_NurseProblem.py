# shift_planner_cp_sat.py
from ortools.sat.python import cp_model
from typing import Dict, List, Tuple, Union
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# --------------------------
# Helper functions
# --------------------------
def intervals_overlap(a: Tuple[float, float], b: Tuple[float, float]) -> bool:
    return max(a[0], b[0]) < min(a[1], b[1])

def shift_intervals_overlap_any(assigned_intervals: List[Tuple[float, float]],
                                forbidden_intervals: List[Tuple[float, float]]) -> bool:
    for a in assigned_intervals:
        for b in forbidden_intervals:
            if intervals_overlap(a, b):
                return True
    return False

def hours_between_intervals(day_a: int, interval_a: Tuple[float, float],
                            day_b: int, interval_b: Tuple[float, float]) -> float:
    end_a = day_a * 24.0 + interval_a[1]
    start_b = day_b * 24.0 + interval_b[0]
    return start_b - end_a

def shifts_violate_rest(day_a: int, intervals_a: List[Tuple[float, float]],
                        day_b: int, intervals_b: List[Tuple[float, float]],
                        min_rest_hours: float) -> bool:
    for ia in intervals_a:
        for ib in intervals_b:
            if hours_between_intervals(day_a, ia, day_b, ib) < min_rest_hours:
                return True
    return False

# --------------------------
# Core function
# --------------------------
def plan_shifts(
    names: List[str],
    shift_requests: List[List[List[int]]],     # shape: [persons][days][4 standard shifts]
    STANDARD_SHIFTS: List[str],                # ordered 4 standard shift names
    shift_time_map: Dict[str, List[Tuple[float, float]]],  # shift_name -> list of (start_hr, end_hr)
    target_shifts: Dict[str, int],
    num_days: int = None,
    shift_requirements: Union[int, Dict[str, Union[int, List[int]]]] = 1,
    min_shifts_per_person: Union[int, List[int]] = 0,
    max_shifts_per_person: Union[int, List[int]] = 999,
    min_rest_hours: float = 8.0,
    time_limit_seconds: int = 20,
    avoid_double_shift_pairs_daywise: List[Tuple[int, str, int, str]] = None,
    double_shift_penalty_weight: int = 50,
    soft_single_shift_weight: int = 20,


):
    P = len(names)
    if num_days is None:
        num_days = len(shift_requests[0])

    # normalize min/max per person
    if isinstance(min_shifts_per_person, int):
        min_shifts = [min_shifts_per_person] * P
    else:
        min_shifts = list(min_shifts_per_person)

    if isinstance(max_shifts_per_person, int):
        max_shifts = [max_shifts_per_person] * P
    else:
        max_shifts = list(max_shifts_per_person)

    # requirement accessor
    def req_for(shift_name: str, day: int) -> int:
        if isinstance(shift_requirements, int):
            return shift_requirements
        if isinstance(shift_requirements, dict):
            v = shift_requirements.get(shift_name, None)
            if v is None:
                return 1
            if isinstance(v, int):
                return v
            else:
                return v[day]
        return 1

    model = cp_model.CpModel()

    # decision vars
    assign = {}
    shift_names = list(shift_time_map.keys())

    for p in range(P):
        for d in range(num_days):
            for s in shift_names:
                assign[(p, d, s)] = model.NewBoolVar(f"assign_p{p}_d{d}_s_{s}")

    # 1) coverage constraints
    for d in range(num_days):
        for s in shift_names:
            req = req_for(s, d)
            model.Add(sum(assign[(p, d, s)] for p in range(P)) == req)

    # 2) forbid assignments marked -1
    std_intervals_map = {std: shift_time_map[std] for std in STANDARD_SHIFTS}
    for p in range(P):
        for d in range(num_days):
            for std_idx, std_name in enumerate(STANDARD_SHIFTS):
                if shift_requests[p][d][std_idx] == -1:
                    forbidden_intervals = std_intervals_map[std_name]
                    for s in shift_names:
                        assigned_intervals = shift_time_map[s]
                        if shift_intervals_overlap_any(assigned_intervals, forbidden_intervals):
                            model.Add(assign[(p, d, s)] == 0)

    # 3) min/max shifts per person
    for p in range(P):
        total = [assign[(p, d, s)] for d in range(num_days) for s in shift_names]
        model.Add(sum(total) >= min_shifts[p])
        model.Add(sum(total) <= max_shifts[p])

    # 4) min rest constraints
    for p in range(P):
        person_shifts = [(d, s, shift_time_map[s]) for d in range(num_days) for s in shift_names]
        N = len(person_shifts)
        for i in range(N):
            d1, s1, intervals1 = person_shifts[i]
            for j in range(i+1, N):
                d2, s2, intervals2 = person_shifts[j]
                if d1 == d2 and s1 == s2:
                    continue
                '''if shifts_violate_rest(d1, intervals1, d2, intervals2, min_rest_hours) \
                   or shifts_violate_rest(d2, intervals2, d1, intervals1, min_rest_hours):
                    model.Add(assign[(p, d1, s1)] + assign[(p, d2, s2)] <= 1)'''
                if shifts_violate_rest(d1, intervals1, d2, intervals2, min_rest_hours):
                    model.Add(assign[(p, d1, s1)] + assign[(p, d2, s2)] <= 1)

    # 5) preferences
    preference_terms = []
    for p in range(P):
        for d in range(num_days):
            for std_idx, std_name in enumerate(STANDARD_SHIFTS):
                if shift_requests[p][d][std_idx] == 1:
                    pref_intervals = shift_time_map[std_name]
                    for s in shift_names:
                        intervals = shift_time_map[s]
                        if shift_intervals_overlap_any(intervals, pref_intervals):
                            preference_terms.append(assign[(p, d, s)])


    # 6) force number of shifts per person
    for p, person_name in enumerate(names):
        if person_name in target_shifts:
            total_shifts = sum(assign[(p, d, s)] for d in range(num_days) for s in shift_names)
            model.Add(total_shifts == target_shifts[person_name])

    # 7) soft penalties: avoid same person assigned to both specified (day, shift) pairs
    double_shift_penalties = []
    if avoid_double_shift_pairs_daywise:
        for (day_a, shift_a, day_b, shift_b) in avoid_double_shift_pairs_daywise:
            for p in range(P):
                both = model.NewBoolVar(f"both_p{p}_d{day_a}_{shift_a}_d{day_b}_{shift_b}")
                model.AddBoolAnd([assign[(p, day_a, shift_a)], assign[(p, day_b, shift_b)]]).OnlyEnforceIf(both)
                model.AddBoolOr([
                    assign[(p, day_a, shift_a)].Not(),
                    assign[(p, day_b, shift_b)].Not()
                ]).OnlyEnforceIf(both.Not())
                double_shift_penalties.append(both)

    # 8) soft penalty: prefer at most one shift per person per day
    multi_shift_penalties = []
    for p in range(P):
        for d in range(num_days):
            shifts_today = [assign[(p, d, s)] for s in shift_names]
            # number of assigned shifts that day
            total_today = sum(shifts_today)
            # penalize each shift beyond the first
            # (since BoolVars, this expression is linear)
            excess_expr = total_today - 1
            # only positive part should be penalized
            # but CP-SAT doesn't have max(), so we just rely on objective minimizing this
            # because excess_expr will be 0 if <=1, >0 if more
            multi_shift_penalties.append(excess_expr)


    # balancing term
    max_shifts_var = model.NewIntVar(0, num_days * len(shift_names), "max_shifts")
    for p in range(P):
        model.Add(sum(assign[(p, d, s)] for d in range(num_days) for s in shift_names) <= max_shifts_var)

    model.Maximize(
        1000 * sum(preference_terms)
        - max_shifts_var
        - double_shift_penalty_weight * sum(double_shift_penalties)
        - soft_single_shift_weight * sum(multi_shift_penalties)
    )

    # solve
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit_seconds
    solver.parameters.num_search_workers = 8
    result = solver.Solve(model)

    if result in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        solution = {}
        shift_counts = {names[p]: 0 for p in range(P)}
        for p in range(P):
            for d in range(num_days):
                for s in shift_names:
                    if solver.Value(assign[(p, d, s)]):
                        solution[(names[p], d, s)] = 1
                        shift_counts[names[p]] += 1

        # ---- check for double-shift occurrences ----
        double_shift_status = []
        if avoid_double_shift_pairs_daywise:
            for (day_a, shift_a, day_b, shift_b) in avoid_double_shift_pairs_daywise:
                violators = []
                for p, person_name in enumerate(names):
                    val_a = solver.Value(assign[(p, day_a, shift_a)])
                    val_b = solver.Value(assign[(p, day_b, shift_b)])
                    if val_a and val_b:
                        violators.append(person_name)
                if violators:
                    double_shift_status.append({
                        "pair": (day_a, shift_a, day_b, shift_b),
                        "violators": violators
                    })
                else:
                    double_shift_status.append({
                        "pair": (day_a, shift_a, day_b, shift_b),
                        "violators": []
                    })

        multi_shift_status = []
        for p, person_name in enumerate(names):
            multi_days = []
            for d in range(num_days):
                shifts_today = [s for s in shift_names if solver.Value(assign[(p, d, s)])]
                if len(shifts_today) > 1:
                    multi_days.append((d, shifts_today))
            if multi_days:
                multi_shift_status.append({"person": person_name, "days": multi_days})

        # Add to return
        return {
            "status": solver.StatusName(result),
            "solution": solution,
            "shift_counts": shift_counts,
            "objective": solver.ObjectiveValue(),
            "double_shift_status": double_shift_status,
            "multi_shift_status": multi_shift_status
        }
    else:
        return None


def check_solution(
    solution: Dict[Tuple[str,int,str], int],
    names: List[str],
    shift_requests: List[List[List[int]]],
    STANDARD_SHIFTS: List[str],
    shift_time_map: Dict[str, List[Tuple[float, float]]],
    shift_requirements,
    min_shifts_per_person=0,
    max_shifts_per_person=999,
    min_rest_hours=8.0
):
    violations = []
    P = len(names)
    name_to_idx = {n: i for i,n in enumerate(names)}
    num_days = len(shift_requests[0])
    shift_names = list(shift_time_map.keys())

    # normalize min/max
    if isinstance(min_shifts_per_person, int):
        min_shifts = [min_shifts_per_person] * P
    else:
        min_shifts = list(min_shifts_per_person)

    if isinstance(max_shifts_per_person, int):
        max_shifts = [max_shifts_per_person] * P
    else:
        max_shifts = list(max_shifts_per_person)

    # requirement accessor
    def req_for(shift_name: str, day: int) -> int:
        if isinstance(shift_requirements, int):
            return shift_requirements
        if isinstance(shift_requirements, dict):
            v = shift_requirements.get(shift_name, None)
            if v is None:
                return 1
            if isinstance(v, int):
                return v
            else:
                return v[day]
        return 1

    # 1) Coverage check
    for d in range(num_days):
        for s in shift_names:
            assigned = [pname for (pname, dd, ss), v in solution.items() if v and dd == d and ss == s]
            req = req_for(s, d)
            if len(assigned) != req:
                violations.append(f"[Coverage] Day {d}, shift {s}: assigned={len(assigned)}, required={req}")

    # 2) Forbidden shifts (-1)
    for (pname, d, s), v in solution.items():
        if v:
            p = name_to_idx[pname]
            for std_idx,std_name in enumerate(STANDARD_SHIFTS):
                if shift_requests[p][d][std_idx] == -1:
                    forbidden_intervals = shift_time_map[std_name]
                    if shift_intervals_overlap_any(shift_time_map[s], forbidden_intervals):
                        violations.append(f"[Forbidden] {pname} assigned to {s} on day {d} but marked -1 for {std_name}")

    # 3) Min/max per person
    for p, pname in enumerate(names):
        total = sum(v for (pp,_,_),v in solution.items() if v and pp==pname)
        if total < min_shifts[p]:
            violations.append(f"[Min shifts] {pname}, has {total}, min {min_shifts[p]}")
        if total > max_shifts[p]:
            violations.append(f"[Max shifts] {pname}, has {total}, max {max_shifts[p]}")

    # 4) Rest violations (with debug info)
    for p, pname in enumerate(names):
        assigned_shifts = [(d,s,shift_time_map[s]) for (pp,d,s),v in solution.items() if v and pp==pname]
        assigned_shifts.sort(key=lambda x:(x[0],x[2][0][0]))
        for i in range(len(assigned_shifts)):
            d1,s1,intervals1 = assigned_shifts[i]
            for j in range(i+1,len(assigned_shifts)):
                d2,s2,intervals2 = assigned_shifts[j]
                # check both directions (since intervals may cross midnight)
                for ia in intervals1:
                    for ib in intervals2:
                        diff = hours_between_intervals(d1, ia, d2, ib)
                        if diff < min_rest_hours:
                            violations.append(
                                f"[Rest] {pname} between {s1}(day{d1}, {ia}) "
                                f"and {s2}(day{d2}, {ib}), rest={diff:.1f}h < {min_rest_hours}h"
                            )
                '''for ia in intervals2:
                    for ib in intervals1:
                        diff = hours_between_intervals(d2, ia, d1, ib)
                        if diff < min_rest_hours:
                            violations.append(
                                f"[Rest] {pname} between {s2}(day{d2}, {ia}) "
                                f"and {s1}(day{d1}, {ib}), rest={diff:.1f}h < {min_rest_hours}h"
                            )'''

    return violations


def build_solution_dict(names: list,
                        days_shifts: list) -> dict:
    solution = {}
    num_days = len(days_shifts)

    for d in range(num_days):
        day = days_shifts[d]
        for shift_name, assigned_people in day.items():
            for person in assigned_people:
                if person not in names:
                    raise ValueError(f"Person {person} not in names list")
                solution[(person, d, shift_name)] = 1
    return solution


def save_solution_to_excel(solution, shift_requirements, num_days, filename="schedule.xlsx"):
    """
    Save the shift assignment solution to an Excel file.

    Each day is a separate sheet.
    Each row is a shift.
    Each assigned person appears in a separate column.
    """

    wb = Workbook()

    for d in range(num_days):
        # Create or select sheet for this day
        if d == 0:
            ws = wb.active
            ws.title = f"Day {d}"
        else:
            ws = wb.create_sheet(title=f"Day {d}")

        # Title
        ws["A1"] = f"Day {d}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        row = 3
        ws.column_dimensions['A'].width = 25
        for c in range(2, 10):
            ws.column_dimensions[chr(64 + c)].width = 20

        for shift_name, req_list in shift_requirements.items():
            # Only include shifts required for this day
            if d < len(req_list) and req_list[d] > 0:
                # Find assigned people
                assigned = [p for (p, dd, s), v in solution.items()
                            if v and dd == d and s == shift_name]

                if assigned:
                    ws.cell(row=row, column=1, value=shift_name).font = Font(bold=True)

                    for i, person in enumerate(assigned, start=2):
                        ws.cell(row=row, column=i, value=person)

                    row += 1

    wb.save(filename)
    print(f"✅ Schedule saved to '{filename}'")

def print_solution_by_day(solution, shift_requirements, num_days):
    """
    Pretty-print the solution using shift_requirements of form:
      {shift_name: [required_day0, required_day1, ...]}
    Only prints shifts with required > 0.
    """
    for d in range(num_days):
        print(f"\nDay {d}")
        for shift_name, req_list in shift_requirements.items():
            if d < len(req_list) and req_list[d] > 0:
                assigned = [p for (p, dd, s), v in solution.items()
                            if v and dd == d and s == shift_name]
                if assigned:  # print only if names assigned
                    print(f"  {shift_name}: {', '.join(assigned)}")


# --------------------------
# Example usage
# --------------------------
if __name__ == "__main__":

    names = ['אברהם דיאמנד', 'אהרון רוטנברג', 'אוהד אלקיים', 'אור חבזה', 'אסף ברזיס', 'דביר רוזנברג', 'דין קרמזין',
             'יאיר בן יוסף', 'יובל ברוכיאן', 'יפתח בן זמרה', 'עודד טובולסקי', 'רונן איזיק', 'רזיאל זקבך']
    shift_requests = [
        [[0, 0, -1, -1], [0, 0, -1, 0], [0, -1, 0, 0], [0, 0, 0, 0], [0, 0, -1, -1], [0, 0, -1, -1], [0, 0, 0, 0]],
        # אברהם דיאמנד
        [[-1, -1, -1, -1], [0, 0, -1, 0], [0, -1, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]],
        # אהרון רוטנברג
        [[-1, 0, 1, 0], [-1, -1, -1, -1], [-1, -1, -1, 0], [0, 0, 1, 0], [0, 0, 1, 0], [0, 0, 1, 0], [0, 1, 0, 0]],
        # אוהד אלקיים
        [[-1, -1, -1, -1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1]],
        # אור חבזה
        [[1, 0, 0, 0], [1, -1, -1, -1], [-1, -1, -1, -1], [1, 0, 0, 0], [1, 0, 0, 0], [1, 0, 0, 0], [1, 0, 0, 0]],
        # אסף ברזיס
        [[0, 1, 0, 0], [0, 1, -1, 0], [-1, -1, -1, -1], [0, 1, 0, 0], [0, 1, -1, 0], [0, 1, 0, 0], [0, 0, 0, 0]],
        # דביר רוזנברג
        [[0, 0, 1, 0], [1, -1, -1, -1], [-1, -1, -1, -1], [0, 0, 1, 0], [0, 0, 1, 0], [0, 0, 1, 0], [0, 0, 0, 0]],
        # דין קרמזין
        [[0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]],
        # יאיר בן יוסף
        [[-1, 0, 1, 0], [1, 0, 0, 0], [1, 0, 0, 0], [-1, -1, 0, 0], [-1, -1, 0, 0], [-1, -1, 1, 0], [0, 0, 0, 0]],
        # יובל ברוכיאן
        [[0, -1, 0, 1], [0, -1, -1, -1], [1, 0, 0, -1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 0]],
        # יפתח בן זמרה
        [[0, 0, -1, -1], [0, 0, 0, 0], [1, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [-1, -1, -1, -1], [-1, 0, 0, -1]],
        # עודד טובולסקי
        [[-1, 0, -1, 0], [1, -1, 0, 0], [0, -1, 0, 0], [0, -1, 0, -1], [-1, -1, 1, 0], [-1, -1, -1, -1], [0, 0, 0, 0]],
        # רונן איזיק
        [[1, 0, 0, 0], [1, 0, 0, 0], [0, 0, 0, 1], [0, 0, 0, 1], [1, 0, 0, 0], [1, 0, 0, 0], [-1, -1, -1, -1]],
        # רזיאל זקבך
    ]



    num_days = 7
    '''cols_to_keep = [0, 1]  # indices of columns you want to keep
    shift_requests = [[row[i] for i in cols_to_keep] for row in shift_requests]'''


    STANDARD_SHIFTS = ["03:00-09:00","09:00-15:00","15:00-21:00","21:00-03:00"]
    shift_time_map = {
        "03:00-09:00": [(3,9)],
        "09:00-15:00": [(9,15)],
        "15:00-21:00": [(15,21)],
        "21:00-03:00": [(21,27)],
        "יזומה 149 בוקר": [(6,9),(17,18.5)],
        "יזומה 149 בוקר ק": [(6,9)],
        "יזומה 149 ערב": [(18.5,23)],
        "דרום 18-24": [(18,24)]
    }


    shift_requirements = {
        "03:00-09:00":          [1, 2, 2, 2, 2, 2, 1 ],  # Day0 → 2 persons, Day1 → 1 person, Day2 → 2 persons
        "09:00-15:00":          [1, 2, 3, 0, 0, 0, 0  ],
        "15:00-21:00":          [1, 2, 2, 0, 1, 0, 1 ],
        "21:00-03:00":          [2, 3, 2, 2, 1, 1, 2  ],
        "יזומה 149 בוקר ק":      [0, 2, 0, 0, 0, 0, 0  ],
        "יזומה 149 בוקר":        [3, 0, 0, 2, 3, 3, 3  ],
        "יזומה 149 ערב":         [3, 0, 0, 3, 3, 3, 2  ],
        "דרום 18-24":            [0, 0, 0, 0, 0, 1, 1  ],
    }



    target_shifts = {
    }

    out = plan_shifts(
        names=names,
        shift_requests=shift_requests,
        STANDARD_SHIFTS=STANDARD_SHIFTS,
        shift_time_map=shift_time_map,
        target_shifts=target_shifts,
        num_days=num_days,
        shift_requirements=shift_requirements,
        min_shifts_per_person=5,
        max_shifts_per_person=6,
        min_rest_hours=12,
        time_limit_seconds=100,
        avoid_double_shift_pairs_daywise=[
            (1, "15:00-21:00", 2, "09:00-15:00")
        ],
        double_shift_penalty_weight=50,
        soft_single_shift_weight = 20,
    )


    if out:
        print("Status:", out["status"])
        print("Objective:", out["objective"])
        print("Shift counts:", out["shift_counts"])
        if "double_shift_status" in out:
            print("\nDouble-shift pair summary:")
            for entry in out["double_shift_status"]:
                day_a, shift_a, day_b, shift_b = entry["pair"]
                violators = entry["violators"]
                if violators:
                    print(
                        f" ⚠️ Pair ({day_a}:{shift_a}) → ({day_b}:{shift_b}) has {len(violators)} violators: {', '.join(violators)}")
                else:
                    print(f" ✅ Pair ({day_a}:{shift_a}) → ({day_b}:{shift_b}) has no violators.")

        if out.get("multi_shift_status"):
            print("\nMultiple-shift-per-day summary:")
            for entry in out["multi_shift_status"]:
                days_info = ", ".join([f"Day {d}: {', '.join(shifts)}" for d, shifts in entry["days"]])
                print(f" ⚠️ {entry['person']} has multiple shifts on {days_info}")
        else:
            print("\n✅ No one has multiple shifts per day.")

        print("Assignments:")
        '''for (p,d,s),v in out["solution"].items():
            print(f" Day {d}, {s} -> {p}")'''
        print_solution_by_day(out["solution"], shift_requirements, num_days)
        save_solution_to_excel(out["solution"], shift_requirements, num_days, filename="schedule.xlsx")
    else:
        print("No feasible solution")
