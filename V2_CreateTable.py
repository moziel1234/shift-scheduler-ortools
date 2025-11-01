import openpyxl
from py._builtin import enumerate

days = 7
shifts = 4
wb = openpyxl.load_workbook(rf"העדפות שמירה.xlsx", data_only=True)
sheet = wb['Sheet1']

names_list = []
min_row = 1
first_name_line_num = -1

# Iterate over the rows in column A
for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
    name = row[0]  # Assuming the name is in the first column (column A)
    min_row += 1
    if name:  # Check if the cell is not empty
        names_list.append(name)
        if first_name_line_num == -1:
            first_name_line_num = min_row - 1

results = []
row_ind = 0
for row in sheet.iter_rows(min_row=first_name_line_num, max_row=first_name_line_num+len(names_list)-1, min_col=2, max_col=days*shifts+1, values_only=True):
    weekly_shifts = list()
    for day in range(days):
        day_shifts = list()
        for shift in range(shifts):
            # if row[day*shifts + shift] == "כן":
            #     day_shifts.append(1)
            if row[day*shifts + shift] is not None and "לא" in row[day*shifts + shift]:
                day_shifts.append(-1)
            elif row[day*shifts + shift] is None:
                day_shifts.append(0)
            else:
                day_shifts.append(1)
                # raise ValueError(rf"Error!!!!! wrong value. row_ind={row_ind}, day={day} val='{row[day*shifts + shift]}'")

        weekly_shifts.append(day_shifts)
    results.append(weekly_shifts)
    row_ind += 1

print("names = ", end="")
print(names_list)

print("shift_requests = [")
for name_ind, name in enumerate(names_list):
    print("[", end="")
    for i, day_shifts in enumerate(results[name_ind]):
        e = ", "
        if i==len(results[name_ind])-1:
            e = ""
        print(day_shifts, end=e)
    print(f"], #{name}")
print("]")