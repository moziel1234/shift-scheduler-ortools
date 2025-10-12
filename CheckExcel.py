from openpyxl import load_workbook

# Load the Excel file
workbook = load_workbook(rf"C:\Users\uziki\OneDrive\שולחן העבודה\2שבצק.xlsx")
sheet = workbook['Sheet3']

# Iterate over each row
i =0
for row in sheet.iter_rows(values_only=True):
    # Get the names from columns A, B, C, D
    names = set(row[0:4])

    # Get the strings from columns E, F, G
    strings = set(row[4:8])

    # Check if any name is substring of any string
    if strings is not None:
        for name in names:
            if name is not None:
                # print(name)
                for string in strings:
                    # print(string)
                    if string is not None:
                        if name in str(string):
                            print(f"row {i} Name '{name}' from columns A, B, C, D is a substring of string '{string}' in columns E, F, G")
    i += 1

    previous_indices = {}

# Iterate over each row
for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    # Get the names from columns A, B, C, D
    names = set(row[0:4])

    # Iterate over each name
    for name in names:
        # Check if the name is already in the dictionary and if its previous occurrence is within 2 rows
        if name is not None:
            if name in previous_indices and i - previous_indices[name] <= 2:
                print(f"Name '{name}' occurs within 2 rows of its previous occurrence.")

        # Update the index of the current occurrence of the name
        previous_indices[name] = i

'''
# Dictionary to store the count of occurrences of each name
name_counts = {}

# Iterate over each row
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Get the names from columns A, B, C, D
    names = set(row[0:4])

    # Update the counts for each name
    for name in names:
        name_counts[name] = name_counts.get(name, 0) + 1

# Print the counts for each name
for name, count in name_counts.items():
    print(f"Name '{name}' appeared {count} times in columns A, B, C, D.")
'''