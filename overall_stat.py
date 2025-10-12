from openpyxl import load_workbook
from collections import defaultdict
import matplotlib.pyplot as plt


def check_string_in_row(row, target_string):
    for i, cell_value in enumerate(row):
        if isinstance(cell_value, str) and target_string == cell_value:
            return i
    return -1

target_name = "עמוס"

# Load the Excel file
workbook = load_workbook(rf"C:\Users\uziki\OneDrive\שולחן העבודה\2שבצק.xlsx")
sheet = workbook['Sheet1']

roles_counter = defaultdict(int)

# Iterate through each row in the sheet
row_ind = 0
shifts = 0
for row in sheet.iter_rows(values_only=True):
    # Assuming the role titles are in rows with red background
    if row and\
            (check_string_in_row(row, "שג")!=-1 or
             check_string_in_row(row, "ש.ג")!=-1  or
             check_string_in_row(row, "ש.ג. 1")!=-1  or
             check_string_in_row(row, "ש.ג.")!=-1):
        role_title = row
    # Assuming names are in the next 4 rows after the role title
    elif target_name in row:
        ind = check_string_in_row(row, target_name)
        role = role_title[ind]
        if not role:
            print(f"{target_name}  {role} ({row_ind},{ind})")
            a=3
        if not role or "שג" in role.replace(".", ""):
            role = "שג"
        roles_counter[role[::-1]] += 1
        shifts +=1
        print(f"{target_name}  {role} ({row_ind},{ind})")
    row_ind+=1

# Plotting the histogram
plt.bar(roles_counter.keys(), roles_counter.values())
plt.xlabel('Roles')
plt.ylabel('Frequency')
plt.title(f'Frequency of {target_name[::-1]} in Different Roles. Total shifts={shifts}')
plt.xticks(rotation=45)
plt.tight_layout()
plt.show()