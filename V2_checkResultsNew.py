import openpyxl

names = ['אביב ברזל', 'אברהם דיאמנד', 'אהרון רוטנברג', 'אוהד אלקיים', 'אור חבזה', 'משה עוזיאל', 'אלעד מאיר',
             'אסף ברזיס', 'דביר רוזנברג', 'דין קרמזין', 'יאיר בן יוסף', 'יאיר מימון', 'יואל שפץ', 'יובל ברוכיאן',
             'יונתן פרידלנדר', 'יפתח בן זמרה', 'מאיר סמסון', 'משה וילנסקי', 'משה ליפן', 'עודד טובולסקי', 'עקיבא עמיאל',
             'רונן איזיק', 'רזיאל זקבך']
shift_requests = [
        [[-1, -1, -1, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, -1, 0], [0, -1, -1, 0], [-1, -1, -1, -1]],  # אביב ברזל
        [[0, 0, -1, -1], [0, -1, -1, 0], [0, 0, 0, 0], [0, 0, -1, 0], [0, -1, 0, 0], [0, 0, 0, 0]],  # אברהם דיאמנד
        [[-1, 0, -1, -1], [-1, -1, -1, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]],  # אהרון רוטנברג
        [[0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]],  # אוהד אלקיים
        [[0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1]],  # אור חבזה
        [[0, 0, 0, 1], [0, 0, 0, 0], [0, 0, -1, 0], [0, 0, 0, 1], [1, 0, 0, 0], [0, 0, 1, 0]],  # משה עוזיאל
        [[0, 0, 0, 0], [0, 0, 0, 0], [-1, -1, -1, -1], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]],  # אלעד מאיר
        [[1, 0, 0, 0], [-1, -1, -1, -1], [1, 0, 0, 0], [1, 0, 0, 0], [-1, -1, -1, -1], [1, 0, 0, 0]],  # אסף ברזיס
        [[-1, 0, 0, 0], [0, 0, 0, 0], [-1, -1, -1, -1], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]],  # דביר רוזנברג
        [[-1, 0, 0, 0], [0, 0, 0, 0], [-1, -1, -1, -1], [1, 0, -1, 0], [0, 0, 0, 0], [1, 0, -1, 1]],  # דין קרמזין
        [[0, 1, 0, 0], [0, 0, 0, 1], [0, 0, 1, 0], [0, 1, 0, 0], [-1, -1, -1, -1], [0, 0, 0, 1]],  # יאיר בן יוסף
        [[-1, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]],  # יאיר מימון
        [[1, 0, -1, -1], [-1, -1, 1, 0], [0, 0, 1, 1], [1, 0, -1, -1], [0, -1, -1, -1], [0, -1, 0, 0]],  # יואל שפץ
        [[-1, 1, -1, -1], [-1, -1, -1, 0], [0, 0, 1, 0], [0, 1, -1, 0], [0, -1, 1, 0], [-1, -1, 1, 0]],  # יובל ברוכיאן
        [[0, 0, 0, 0], [0, 0, 1, 0], [-1, -1, -1, -1], [0, 0, 0, 0], [0, 0, 0, 0], [1, 0, 0, 0]],  # יונתן פרידלנדר
        [[0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1], [0, 0, 0, 1]],  # יפתח בן זמרה
        [[0, 1, 0, -1], [0, 0, 0, 1], [0, -1, 1, 0], [0, 0, 0, 1], [0, 0, 0, 0], [0, 0, 0, 0]],  # מאיר סמסון
        [[0, 1, -1, -1], [-1, -1, -1, -1], [0, 0, 1, 0], [0, 0, -1, -1], [0, -1, 1, 0], [0, 1, 0, 0]],  # משה וילנסקי
        [[1, -1, -1, 0], [1, 0, 0, 0], [1, 0, 0, 0], [1, 0, 0, 0], [1, 0, 0, 0], [1, 0, 0, 0]],  # משה ליפן
        [[0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [1, 0, 0, 0], [0, 0, 0, 0]],  # עודד טובולסקי
        [[0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]],  # עקיבא עמיאל
        [[-1, 0, 0, 0], [0, 0, 0, 0], [-1, -1, -1, -1], [0, 0, 0, 0], [0, -1, 0, 0], [0, -1, -1, -1]],  # רונן איזיק
        [[1, 0, 0, -1], [0, -1, 0, 0], [0, -1, -1, -1], [-1, 0, 0, -1], [0, 0, 0, 1], [0, 0, 0, 0]],  # רזיאל זקבך
    ]

# standard-shift names and their index in shift_requests[person][day][index]
STANDARD_SHIFTS = ["03:00-09:00", "09:00-15:00", "15:00-21:00", "21:00-03:00"]

# Map shift titles -> time ranges
shift_time_map = {
    "03:00-09:00": [(3, 9)],
    "09:00-15:00": [(9, 15)],
    "15:00-21:00": [(15, 21)],
    "21:00-03:00": [(21, 27)],
    "יזומה 149 בוקר": [(6, 9), (17, 18.5)],
    "יזומה 149 בוקר ק": [(6, 9)],
    "יזומה 149 ערב": [(18.5, 23)],
    "דרום 18-24": [(18, 24)],
    "הורדיון 09-13": [(9, 13)],
    "הורדיון 13-17": [(13, 17)],
}





def intervals_overlap(a, b):
    """Return True when intervals a=(s1,e1) and b=(s2,e2) overlap strictly."""
    return max(a[0], b[0]) < min(a[1], b[1])

def violates_constraints(person, day_idx, assigned_shift_title, names, shift_requests):
    """
    Check if `person` assigned to `assigned_shift_title` on day `day_idx`
    violates any '-1' (can't) entry in shift_requests.

    Returns: (bool_is_violation, reason_string_or_None)
    """
    if not person:
        return False, None
    person = person.strip()
    try:
        p_idx = names.index(person)
    except ValueError:
        return False, f"person '{person}' not in names list"

    # get intervals for the assigned shift
    assigned_intervals = shift_time_map.get(assigned_shift_title)
    if not assigned_intervals:
        return False, f"no time mapping for shift '{assigned_shift_title}'"

    # safety: ensure day_idx in range for this person
    if day_idx < 0 or day_idx >= len(shift_requests[p_idx]):
        return False, f"day index {day_idx} out of range for person '{person}'"

    # for each standard shift that person marked -1 that day, check interval overlap
    for std_idx, std_name in enumerate(STANDARD_SHIFTS):
        if shift_requests[p_idx][day_idx][std_idx] == -1:
            forbidden_intervals = shift_time_map.get(std_name)
            if not forbidden_intervals:
                # This should not happen for standard shifts, but skip if missing
                continue
            for a in assigned_intervals:
                for b in forbidden_intervals:
                    if intervals_overlap(a, b):
                        reason = (f"'{person}' assigned to '{assigned_shift_title}' (interval {a}) "
                                  f"overlaps forbidden standard shift '{std_name}' (interval {b}) on day {day_idx}")
                        return True, reason

    return False, None
# Track assignments
violations = []
assignments = {name: [] for name in names}
shift_counts = {name: 0 for name in names}

wb = openpyxl.load_workbook("plan.xlsx")
ws = wb.active

# Assuming each day block has 4 columns
num_days = (ws.max_column) // 4

for d in range(num_days):
    col_start = d * 4 + 1
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_start, max_col=col_start+3):
        shift_title = row[0].value
        if not shift_title:
            continue
        if shift_title not in shift_time_map:
            continue

        for cell in row[1:]:
            person = cell.value
            if not person:
                continue
            # print(person)

            # Save assignment
            shift_counts[person] += 1
            for interval in shift_time_map[shift_title]:
                assignments[person].append((d, interval, shift_title))

            # Check constraint violation
            if person in names:
                i = names.index(person)
                shift_index = list(shift_time_map.keys()).index(shift_title)
                # Only check -1 constraints for standard shifts (4/day)
                is_violation, reason = violates_constraints(person, d, shift_title, names, shift_requests)
                if is_violation:
                    violations.append((person.strip(), d, shift_title, reason))
                '''if shift_title in ["03:00-09:00", "09:00-15:00", "15:00-21:00", "21:00-03:00"]:
                    if shift_requests[i][d][shift_index] == -1:
                        violations.append((person, d, shift_title))'''

# Check rest time >= 8h
rest_violations = []
min_dist = 1000
min_dist_person = "non"

for person, shifts in assignments.items():
    shifts_sorted = sorted(shifts, key=lambda x: (x[0], x[1][0]))  # sort by day, start time
    for i in range(1, len(shifts_sorted)):
        day1, (s1, e1), shift1 = shifts_sorted[i-1]
        day2, (s2, e2), shift2 = shifts_sorted[i]

        # 🚫 skip if same logical shift (like two intervals of יזומה 149 בוקר)
        if shift1 == shift2:
            continue

        hours_between = (day2*24 + s2) - (day1*24 + e1)
        if hours_between < min_dist:
            min_dist = hours_between
            min_dist_person = person
        if hours_between < 8:
            rest_violations.append((person, shifts_sorted[i-1], shifts_sorted[i]))


print("🚨 Constraint Violations:")
for v in violations:
    print(v)

print("\n⏰ Rest Violations (<8h):")
for rv in rest_violations:
    print(rv)

print("\n📊 Shift Counts:")
for k,v in shift_counts.items():
    print(k, v)

print(min_dist, min_dist_person)

