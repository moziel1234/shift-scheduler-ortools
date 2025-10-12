from openpyxl import load_workbook

# Load the Excel file
workbook = load_workbook(rf"C:\Users\uziki\OneDrive\שולחן העבודה\2שבצק.xlsx")
sheet = workbook['Sheet4']

# List of names
names = [ "ויקטור", "מני", "גבי", "אריה", "יונתן", "מימון", "אפשטיין", "יעקב",
              "עמר", "עקיבא", "משה", "אור", "רזיאל",  "אלון", "אסף", "ישראל"]

# List to store the final result
result = []

# create results names list
for i in range(len(names)):
    result.append([])

# Iterate through each row in the Excel sheet
day = -1
for row in range(1, sheet.max_row + 1):
    # create new day
    check_row = False
    if sheet.cell(row=row, column=1).value == "משמרת 1":
        for i in range(len(names)):
            result[i].append([0, 0, 0, 0])
        day += 1
        check_row = True
        shift = 0
    elif sheet.cell(row=row, column=1).value == "משמרת 2":
        check_row = True
        shift = 1
    elif sheet.cell(row=row, column=1).value == "משמרת 3":
        check_row = True
        shift = 2
    elif sheet.cell(row=row, column=1).value == "משמרת 4":
        check_row = True
        shift = 3

    if check_row:
        # Iterate through each cell in the row
        for col in range(2, 15):  # Assuming negative  names are in columns B and M
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                for name in names:
                    if name in cell_value:
                        result[names.index(name)][day][shift] = -1

        for col in range(15, 21):  # Assuming positive  names are in columns B and M
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                for name in names:
                    if name in cell_value:
                        result[names.index(name)][day][shift] = 1

print("shift_requests = [")
for name in names:
    print("[", end="")
    for i, day_shifts in enumerate(result[names.index(name)]):
        e = ", "
        if i==len(result[names.index(name)])-1:
            e = ""
        print(day_shifts, end=e)
    print(f"], #{name}")
print("]")

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=13):
    # Check if column A contains the string "משמרת"
    if  row[0].value and "משמרת" in row[0].value:
        # Join values from columns B-G with "+"
        joined_values_1 = '+'.join(str(cell.value) for cell in row[1:7] if cell.value)

        # Join values from columns H-M with "+"
        joined_values_2 = '+'.join(str(cell.value) for cell in row[7:13] if cell.value)

        print(joined_values_1)
        # Save the joined values in columns T and U respectively
        sheet.cell(row=row[0].row, column=21).value = joined_values_1  # Column T
        sheet.cell(row=row[0].row, column=22).value = joined_values_2  # Column U

# Save the workbook
workbook.save(r"C:\Users\uziki\OneDrive\שולחן העבודה\2שבצק.xlsx")